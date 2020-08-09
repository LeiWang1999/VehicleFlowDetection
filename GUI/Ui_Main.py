import sys
from PyQt5.QtWidgets import (QMainWindow,
                            QFileDialog,
                            QGraphicsPixmapItem,
                            QGraphicsScene,
                            QLabel,
                            QGridLayout,
                            QHeaderView,
                            QTableWidgetItem,
                            )
from PyQt5 import QtGui
from PyQt5.QtCore import Qt,QThread, pyqtSignal
import threading
from .Ui_VechicleGUI import Ui_MainWindow
import cv2
import numpy as np
import random
import time
from PIL import Image
import tensorflow as tf
from pathlib import Path
import sys
import os
from tqdm import tqdm
import pickle
from openpyxl import Workbook
from openpyxl.styles import PatternFill
lib_dir = (Path(__file__).parent / '..' / 'lib').resolve()
if str(lib_dir) not in sys.path:
    sys.path.insert(0, str(lib_dir))

from core import utils
import tools.save_image as save_image
from tools.trackers_to_perframe import draw_bbox_with_counting, rgbarray2str,visdrone_class_name
from tools.iou_tracker import save_mod, track_viou_video, save_to_csv

class UiMain(QMainWindow):

    def __init__(self):
        super().__init__()
        ui = Ui_MainWindow()
        ui.setupUi(self)
        self.ui = ui
        self.media_path = ""
        # Default Show Realtime Video
        self.showVideo_flag = True
        self.ui.realtimemode.setChecked(True)
        self.writeVideo_flag = False
        self.return_elements = ["input/input_data:0", "pred_sbbox/concat_2:0",
                                "pred_mbbox/concat_2:0", "pred_lbbox/concat_2:0"]
        self.pb_file = "./models/yolov3_visdrone.pb"
        self.output_path = './output/output.avi'
        self.counting_path = './output/counting.avi'
        self.annotation_path = './output/tracker.txt'
        self.pickle_file_path = './output/tmp.pk'
        self.excel_path = './output/'
        self.excel_name = ''
        self.num_classes = 12
        self.input_size = 416
        self.video_time = 0
        self.current_time = 0
        self.is_on = False
        self.LINE_LENTH=ui.baseline_table.rowCount()
        self.start_point = ()
        self.end_point = ()
        self.line_list = []
        self.ws_list = []
        self.draw_flag = 0  # when line start flag = 0 , line end flag = 1

    
        ui.graphicsView.mousePressEvent = self.grapic_view_mousePressEvent
        ui.browse.clicked.connect(self.browse_file)
        ui.start.clicked.connect(self.start_process)
        ui.pause.setEnabled(False)
        ui.pause.clicked.connect(self.pause_process)
        ui.realtimemode.clicked.connect(self.update_realtimemode)
        ui.line_clear.clicked.connect(self.clear_baseline)


    def browse_file(self):
        media_path, media_type = QFileDialog.getOpenFileName(
            self, "Open Media")
        if media_path == "":
            return
        self.media_path = media_path
        # Valiate media
        vid = cv2.VideoCapture(media_path)
        self.vid = vid
        if not vid.isOpened():
            self.ui.textBrowser.setPlainText("Couldn't open media!")
        else:
            self.ui.textBrowser.setPlainText(media_path)
            self.ui.textBrowser.moveCursor(
                self.ui.textBrowser.textCursor().End)
        # Save media info
        self.total_frame_counter = int(vid.get(cv2.CAP_PROP_FRAME_COUNT))
        self.media_fps = vid.get(cv2.CAP_PROP_FPS)
        self.video_time = self.total_frame_counter / self.media_fps
        self.media_size = (int(vid.get(cv2.CAP_PROP_FRAME_WIDTH)),
                           int(vid.get(cv2.CAP_PROP_FRAME_HEIGHT)))
        self.excel_name =  time.strftime("%Y-%m-%d-%H:%M:%S", time.localtime()) + '.xlsx'
        self.excel_path = self.excel_path + self.excel_name
        # Init Excel Workspace
        self.wb = Workbook()
        # Sample frame for baseline
        while True:
            return_value, frame = vid.read()
            if return_value == True:
                frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                self.frame = frame
                break
        self.update_graphic_viewer(frame)

    def update_baseline(self):
        # update baseline viewer
        image = self.frame.copy()
        for line in self.line_list:
            cv2.line(image,
                 line["start_point"],
                 line["end_point"],
                 line["line_color"], 5)
        self.update_graphic_viewer(image)

        # update baseline info
        # 设置一个标签
        for index, line in enumerate(self.line_list):
            Labelindex = QTableWidgetItem(str(index + 1))
            Labelindex.setFont(QtGui.QFont('Times', 14))
            Labelindex.setForeground(QtGui.QBrush(QtGui.QColor(*line["line_color"])))
            Labelindex.setTextAlignment(Qt.AlignCenter)
            LabelStart = QTableWidgetItem(str(line["start_point"]))
            LabelStart.setFont(QtGui.QFont('Times', 14))
            LabelStart.setForeground(QtGui.QBrush(QtGui.QColor(*line["line_color"])))
            LabelStart.setTextAlignment(Qt.AlignCenter)
            LabelEnd = QTableWidgetItem(str(line["end_point"]))
            LabelEnd.setFont(QtGui.QFont('Times', 14))
            LabelEnd.setForeground(QtGui.QBrush(QtGui.QColor(*line["line_color"])))
            LabelEnd.setTextAlignment(Qt.AlignCenter)
            LabelCount = QTableWidgetItem(str(line["line_counter"]))
            LabelCount.setFont(QtGui.QFont('Times', 14))
            LabelCount.setForeground(QtGui.QBrush(QtGui.QColor(*line["line_color"])))
            LabelCount.setTextAlignment(Qt.AlignCenter)
            
            self.ui.baseline_table.setItem(index + 1, 0, Labelindex)
            self.ui.baseline_table.setItem(index + 1, 1, LabelStart)
            self.ui.baseline_table.setItem(index + 1, 2, LabelEnd)
            self.ui.baseline_table.setItem(index + 1, 3, LabelCount)
            

    def update_graphic_viewer(self, image):
        showImage = QtGui.QImage(
            image, image.shape[1], image.shape[0], QtGui.QImage.Format_RGB888)
        pix = QtGui.QPixmap.fromImage(showImage)
        item = QGraphicsPixmapItem(pix)  # 创建像素图元
        self.scene = QGraphicsScene()  # 创建场景
        self.scene.addItem(item)
        self.ui.graphicsView.setScene(self.scene)  # 将场景添加至视图

    def update_baseline_count(self):
        for index, line in enumerate(self.line_list):
            LabelCount = QTableWidgetItem(str(line["line_counter"]))
            LabelCount.setFont(QtGui.QFont('Times', 14))
            LabelCount.setForeground(QtGui.QBrush(QtGui.QColor(*line["line_color"])))
            LabelCount.setTextAlignment(Qt.AlignCenter)
            self.ui.baseline_table.setItem(index + 1, 3, LabelCount)

    '''
        Control mutual Status
    '''

    def mutual_control(self, status: bool):
        self.ui.browse.setEnabled(status)
        self.ui.line_clear.setEnabled(status)
        self.ui.start.setEnabled(status)
        self.ui.pause.setEnabled(status)
        self.ui.realtimemode.setEnabled(status)

    def excel_initial(self):
        for index, line in enumerate(self.line_list):
            self.ws_list.append(self.wb.create_sheet(title='line_' + str(index), index=index))
        for index, ws in enumerate(self.ws_list):
            # init header
            ws.cell(row=1, column=1).value = '类别'
            ws.cell(row=1, column=2).value = '出现时间'
            ws.cell(row=1, column=3).value = '出现次数'
            ws.cell(row=1, column=4).value = '车流量'
            ws.cell(row=1, column=1).fill = PatternFill(fill_type='solid', fgColor=rgbarray2str(self.line_list[index]['line_color']))
            ws.cell(row=1, column=2).fill = PatternFill(fill_type='solid', fgColor=rgbarray2str(self.line_list[index]['line_color']))
            ws.cell(row=1, column=3).fill = PatternFill(fill_type='solid', fgColor=rgbarray2str(self.line_list[index]['line_color']))
            ws.cell(row=1, column=4).fill = PatternFill(fill_type='solid', fgColor=rgbarray2str(self.line_list[index]['line_color']))
            for index, class_name in enumerate(visdrone_class_name):
                ws.cell(row=index + 2, column=1).value = class_name
                ws.cell(row=index + 2, column=2).value = ''
                ws.cell(row=index + 2, column=3).value = 0
                ws.cell(row=index + 2, column=4).value = 0
                                

    '''
        开始进行图像处理操作
    '''
    def start_process(self):
        if len(self.line_list) == 0:
            self.ui.baseline_message.setText("Please Draw Base Line First!!!!")
            return
        # Init excel sheet for each line
        self.excel_initial()
        self.mutual_control(False)
        self.ui.pause.setEnabled(True)
        self.compute_thread = WorkThread(window=self)
        self.compute_thread.update_graphic_viewer.connect(self.update_graphic_viewer)
        self.compute_thread.update_process_message.connect(self.update_process_message)
        self.compute_thread.update_baseline_count.connect(self.update_baseline_count)
        self.compute_thread.start()



    def pause_process(self):
        self.is_on = not self.is_on
        if self.is_on == True:
            self.ui.pause.setText('pause')
        else:
            self.ui.pause.setText('continue')


    def update_realtimemode(self):
        self.showVideo_flag = self.ui.realtimemode.isChecked()
    
    def update_process_message(self, message):
        self.ui.process_message.setText(message)
    
    def grapic_view_mousePressEvent(self, e):
        # Validate
        if self.media_path == "":
            self.ui.baseline_message.setText("Please open media first!")
            return
        if len(self.line_list) >= self.LINE_LENTH:
            self.ui.baseline_message.setText("Baseline Maximized!")
            return
        x = e.x()
        y = e.y()
        scroll_x = self.ui.graphicsView.horizontalScrollBar().sliderPosition()
        scroll_y = self.ui.graphicsView.verticalScrollBar().sliderPosition()
        point_x, point_y = x + scroll_x, y + scroll_y
        if self.draw_flag == 0:
            self.start_point = (point_x, point_y)
            self.draw_flag = 1
        elif self.draw_flag == 1:
            self.end_point = (point_x, point_y)
            line_color = (random.randint(0,255),random.randint(0,255),random.randint(0,255))
            self.line_list.append({"start_point":self.start_point,"end_point":self.end_point,"line_color":line_color, "line_counter": 0})
            self.update_baseline()
            self.draw_flag = 0
    
    def clear_baseline(self):
        self.line_list = []
        self.draw_flag = 0
        self.update_baseline()
        # clear table view
        self.ui.baseline_table.clearContents()
        return


class WorkThread(QThread):
    update_graphic_viewer = pyqtSignal(np.ndarray)
    update_process_message = pyqtSignal(str)
    update_baseline_count = pyqtSignal()
    def __init__(self, window, parent=None):
        super(WorkThread, self).__init__(parent)
        self.window = window
        self.window.is_on = True
    
    def detect_inference(self):
        self.update_process_message.emit('Detection Process')
        graph = tf.Graph()
        detections = []
        return_tensors = utils.read_pb_return_tensors(
            graph, self.window.pb_file, self.window.return_elements)
        # 创建进度条
        pbar = tqdm(total=self.window.total_frame_counter)
        with tf.Session(graph=graph) as sess:
            if self.window.writeVideo_flag:
                isOutput = True if self.window.output_path != "" else False
                if isOutput:
                    video_FourCC = cv2.VideoWriter_fourcc(*'MPEG')
                    out = cv2.VideoWriter(
                        self.window.output_path, video_FourCC, self.window.media_fps, self.window.media_size)
                    frame_index = -1
            while True:
                while not self.window.is_on:
                    pass
                return_value, frame = self.window.vid.read()
                if return_value != True:
                    break
                if return_value:
                    image = Image.fromarray(frame)
                    self.window.frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                else:
                    raise ValueError("No image!")
                frame_size = frame.shape[:2]
                image_data = utils.image_preporcess(
                    np.copy(frame), [self.window.input_size, self.window.input_size])
                image_data = image_data[np.newaxis, ...]


                pred_sbbox, pred_mbbox, pred_lbbox = sess.run(
                    [return_tensors[1], return_tensors[2], return_tensors[3]],
                    feed_dict={return_tensors[0]: image_data})
                pred_bbox = np.concatenate([np.reshape(pred_sbbox, (-1, 5 + self.window.num_classes)),
                                    np.reshape(pred_mbbox, (-1, 5 + self.window.num_classes)),
                                    np.reshape(pred_lbbox, (-1, 5 + self.window.num_classes))], axis=0)
                bboxes = utils.postprocess_boxes(pred_bbox, frame_size, self.window.input_size, 0.45)
                bboxes = utils.nms(bboxes, 0.45, method='nms')
                image = utils.draw_bbox(frame, bboxes)
                # image = utils.draw_bbox(frame, bboxes, vid.get(1))
                # 保存为 iou_tracker 格式
                detections = save_mod(bboxes, 0.6)
                result = np.asarray(image)
                if self.window.writeVideo_flag:
                    # save a frame
                    out.write(result)
                if self.window.showVideo_flag:
                    self.update_graphic_viewer.emit(result)
                    pbar.update(1)
                    self.window.ui.processrate.setText(str(pbar))
                else:
                    pbar.update(1)
                    self.window.ui.processrate.setText(str(pbar))
        # Release everything if job is finished
        if self.window.writeVideo_flag:
            out.release()
        pbar.close()
        # 多目标追踪
        trackers = track_viou_video(self.window.media_path ,detections , 0.5, 0.6, 0.1, 23, 16, 'MEDIANFLOW', 1.0, self.window)
        # 保存 trackers
        with open(self.window.pickle_file_path, 'wb') as pk_f:
            pickle.dump(trackers, pk_f)
            self.window.ui.processrate.setText('=> saved trackers to pk file.')
    
    def vehicle_counting(self):
        self.update_process_message.emit('Counting Process')
        # reopen media capture
        vid = cv2.VideoCapture(self.window.media_path)
        with open(self.window.pickle_file_path, 'rb') as pk_f:
            trackers = pickle.load(pk_f)
            self.window.ui.processrate.setText('=> load trackers from pk file .')
        if self.window.writeVideo_flag:
            isOutput = True if self.window.output_path != "" else False
            if isOutput:
                video_FourCC = cv2.VideoWriter_fourcc(*'MPEG')
                out = cv2.VideoWriter(
                    self.window.counting_path, video_FourCC, self.window.media_fps, self.window.media_size)
        pbar = tqdm(total=self.window.total_frame_counter)

        while True:
            while not self.window.is_on:
                pass
            return_value, frame = vid.read()
            if return_value != True:
                break
            if return_value:
                self.window.frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                pass
            else:
                raise ValueError("No image!")
            self.window.current_time = vid.get(cv2.CAP_PROP_POS_MSEC)
            result = draw_bbox_with_counting(frame, vid.get(1), trackers, self.window)
            self.update_baseline_count.emit()
            if self.window.showVideo_flag:
                pbar.update(1)
                self.window.ui.processrate.setText(str(pbar))
                self.update_graphic_viewer.emit(result)
            else:
                pbar.update(1)
                self.window.ui.processrate.setText(str(pbar))
            if self.window.writeVideo_flag:
                # save a frame
                out.write(result)
        if self.window.writeVideo_flag:
            out.release()
        pbar.close()
        self.window.mutual_control(True)
        self.window.ui.pause.setEnabled(False)

    def __del__(self):
        self.window.is_on = False
        self.wait()

    def run(self):
        self.window.is_on = True
        # main process
        self.detect_inference()
        self.vehicle_counting()
        # save excel
        for index, ws in enumerate(self.window.ws_list):
            # init header
            count_sum = 0
            density_sum = 0
            for index, class_name in enumerate(visdrone_class_name):
                count_sum += ws.cell(row=index + 2, column=3).value
                density_sum += ws.cell(row=index + 2, column=4).value
            
            ws.cell(row=len(visdrone_class_name) + 2, column=1).value = '总和'
            ws.cell(row=len(visdrone_class_name) + 2, column=3).value = count_sum
            ws.cell(row=len(visdrone_class_name) + 2, column=4).value = density_sum
            
            
        self.window.wb.remove(self.window.wb['Sheet'])
        self.window.wb.save(self.window.excel_path)