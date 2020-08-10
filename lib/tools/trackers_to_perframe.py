import cv2
import pdb
from openpyxl import Workbook


visdrone_class_name = ['ignored-regions','pedestrian','people','bicycle','car','van','truck','tricycle','awning-tricycle','bus','motor','others']
headers = ['class', 'time', 'count', 'density']
rows  = [
        {'class':'ignored-regions','time':'','speed':'','count':0,'density':0},
        {'class':'pedestrian','time':'','speed':'','count':0,'density':0},
        {'class':'people','time':'','speed':'','count':0,'density':0},
        {'class':'bicycle', 'time': '', 'speed':'','count':0, 'density': 0},
        {'class':'car','time':'','speed':'','count':0,'density':0},
        {'class':'van','time':'','speed':'','count':0,'density':0},
        {'class':'truck','time':'','speed':'','count':0,'density':0},
        {'class':'tricycle', 'time': '', 'speed':'','count':0,  'density': 0},
        {'class':'awning-tricycle','time':'','speed':'','count':0,'density':0},
        {'class':'bus','time':'','speed':'','count':0,'density':0},
        {'class':'motor','time':'','speed':'','count':0,'density':0},
        {'class':'others','time':'','speed':'','count':0,'density':0},
    ]

bbox_color = {'ignored-regions':(0xFF,0x66,0x00),
    'pedestrian':(0xCC,0x66,0x00),
    'people':(0x99,0x66,0x00),
    'bicycle':(0x66,0x66,0x00),
    'car':(0x33,0xFF,0x00),
    'van':(0x00,0x66,0x00),
    'truck':(0xFF,0xFF,0x00),
    'tricycle':(0xCC,0xFF,0x00),
    'awning-tricycle':(0x99,0xFF,0x00),
    'bus':(0x66,0xFF,0x00),
    'motor':(0x33,0x66,0x00),
    'others':(0x00,0xFF,0x00)}

def draw_bbox_with_counting(image, image_index, trackers, window, show_box=True, show_label=True):
    line_list = window.line_list
    for object_id, object_info in enumerate(trackers):
        if image_index >= object_info['start_frame'] and image_index < (object_info['start_frame'] + len(object_info['bboxes'])):
            object_info_index = int(image_index - object_info['start_frame'])
            bbox = object_info['bboxes'][object_info_index]
            c1 = (int(bbox[0]), int(bbox[1]))
            c2 = (int(bbox[2]), int(bbox[3]))
            speed = bbox[4]
            if show_box:  
                cv2.rectangle(image, c1, c2, bbox_color[object_info['class']], 2)
            if show_label:
                bbox_mess = '%s: %.2f' % (object_info['class'], object_info['max_score'])
                cv2.putText(image, bbox_mess, (c1[0], c1[1]-2), cv2.FONT_HERSHEY_SIMPLEX,
                        0.5, (0, 0, 0), 1, lineType=cv2.LINE_AA)
            # 获取判断列表
            target_object = {
                'ignored-regions': False,
                'pedestrian': window.ui.pedestrian.isChecked(),
                'people':  window.ui.people.isChecked(),
                'bicycle':  window.ui.bicycle.isChecked(),
                'car':  window.ui.car.isChecked(),
                'van':  window.ui.van.isChecked(),
                'truck':  window.ui.truck.isChecked(),
                'tricycle':  window.ui.tricycle.isChecked(),
                'awning-tricycle':  window.ui.awning_tricycle.isChecked(),
                'bus':  window.ui.bus.isChecked(),
                'motor':  window.ui.motor.isChecked(),
                'others': False}
            target_object = list(target_object.keys())[list(target_object.values()).index(True)]
            # 判断
            if object_info['class'] in target_object:       
                if object_info_index > 0:
                    bbox_last = object_info['bboxes'][object_info_index-1]
                    center_now = (((bbox[0])+(bbox[2]))/2, ((bbox[1])+(bbox[3]))/2)
                    center_last = (((bbox_last[0]) + (bbox_last[2])) / 2, ((bbox_last[1]) + (bbox_last[3])) / 2)
                    for index, line in enumerate(line_list):
                        start_point = line["start_point"]
                        end_point = line["end_point"]
                        line_color = line["line_color"]
                        counter = line["line_counter"]
                        vehicle_is_online = counting(center_now, center_last,start_point, end_point)
                        if vehicle_is_online:
                            line["line_counter"] = counter + 1
                            # update excel output
                            # get index and update time
                            ws = window.ws_list[index]
                            index = visdrone_class_name.index(object_info['class'])
                            time = window.current_time / 1000  # 1000ms = 1s
                            rows[index]['time'] = rows[index]['time'] + str(object_id)+ ':' + str(round(time, 2)) + 's,'
                            rows[index]['count'] = line["line_counter"]
                            rows[index]['speed'] = rows[index]['speed'] + str(object_id)+ ':' + str(round(speed, 2)) + 'm/s'
                            rows[index]['density'] = line["line_counter"] / time
                            for index, row in enumerate(rows):
                                ws.cell(row=index + 2, column=1).value = row['class']
                                ws.cell(row=index + 2, column=2).value = row['time']
                                ws.cell(row=index + 2, column=3).value = row['speed']
                                ws.cell(row=index + 2, column=4).value = row['count']
                                ws.cell(row=index + 2, column=5).value = row['density']
                                

                            
    for line in line_list:
        cv2.line(image,
                line["start_point"],
                line["end_point"],
                line["line_color"],
                5)
        start_point = line["start_point"]
        end_point = line["end_point"]
        line_color = line["line_color"]
        info = str(line["line_counter"])
        cv2.putText(image, text=info, 
                org=(start_point[0] + 10, start_point[1] - 20), 
                fontFace=cv2.FONT_HERSHEY_SIMPLEX,
                fontScale=1, color=line_color, thickness=2)
    return image

def cross(p1,p2,p3):
    x1=p2[0]-p1[0]
    y1=p2[1]-p1[1]
    x2=p3[0]-p1[0]
    y2=p3[1]-p1[1]
    return x1 * y2 - x2 * y1

def counting(center_now, center_last, start_point, end_point):
    vehicle_is_online = False
    if (max(center_now[0], center_last[0]) >= min(start_point[0], end_point[0]) \
        and max(start_point[0], end_point[0]) >= min(center_now[0], center_last[0]) \
            and max(center_now[1], center_last[1]) >= min(start_point[1], end_point[1]) \
                and max(start_point[1], end_point[1]) >= min(center_now[1], center_last[1])):
                if (cross(center_now, center_last, start_point) * cross(center_now, center_last, end_point) <= 0 \
                    and cross(start_point, end_point, center_now) * cross(start_point, end_point, center_last) <= 0):
                    vehicle_is_online = True
    return vehicle_is_online   
            
def rgbarray2str(array):
    r = array[0]
    g = array[1]
    b = array[2]
    print(str(hex(r)[2:]).zfill(2)+str(hex(g)[2:]).zfill(2)+str(hex(b)[2:]).zfill(2))
    return str(hex(r)[2:]).zfill(2)+str(hex(g)[2:]).zfill(2)+str(hex(b)[2:]).zfill(2)